"""
Writes the pilot dataset as a real .xlsx (not CSV) so Excel doesn't mangle
it on open: CSV has no type information, so Excel's import heuristics strip
leading zeros from stock codes (e.g. MiniMax's "00100" becomes "100") and
apply no consistent number formatting. An .xlsx cell carries an explicit
type (string vs number) and format, so both are fixed at the source.
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

NUMBER_FORMAT = "#,##0.00"
TEXT_COLUMNS = {"stock_code", "a_share_stock_code", "adr_code"}


def write_xlsx(rows: list[dict], column_names: list[str], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "IPO Data"

    ws.append(column_names)
    for r in rows:
        ws.append([r.get(c) for c in column_names])

    for col_idx, col_name in enumerate(column_names, start=1):
        letter = get_column_letter(col_idx)
        if col_name in TEXT_COLUMNS:
            # force Text format so leading zeros (e.g. stock code "00100")
            # survive -- otherwise Excel re-infers the column as numeric on
            # open and strips them, even though the underlying value is a
            # Python str already.
            for row_idx in range(2, len(rows) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value)
            continue
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            # only apply the money/count format to actual numeric values --
            # bool is a subclass of int in Python, so it's excluded
            # explicitly (True/False must stay as booleans, not "1.00").
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = NUMBER_FORMAT

    wb.save(out_path)
